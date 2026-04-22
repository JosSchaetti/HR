"""
import_names.py — Mitarbeiter- und Lohndaten in glatec_hr.db importieren
=========================================================================
Erwartet die Datei import_template.xlsx (oder beliebige .xlsx/.csv)

Sheet "1_Mitarbeiter":
    erp_id*, first_name, last_name, gender, birth_date,
    department_en, cost_center, position_en, entry_date, fte, status

Sheet "2_Lohndaten":
    erp_id*, band_code, surcharge, actual_net

Neue Mitarbeiter (erp_id nicht in DB) werden angelegt.
Bestehende Mitarbeiter werden aktualisiert (nur befüllte Felder).

Ausführen:
    python import_names.py import_template.xlsx
    python import_names.py import_template.xlsx --dry-run
"""

import sys, os, sqlite3
from datetime import datetime, date

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "glatec_hr.db")

DEPT_BG_BY_EN = {
    'Accounting': 'Счетоводство', 'Assembly': 'Монтаж',
    'Automation': 'Автоматизация', 'Automation Design': 'Дизайн на автоматизация',
    'Costing / Calculation': 'Калкулация', 'Facilities': 'Сгради',
    'General IT': 'Общи IT', 'HR Management': 'Управление ЧР',
    'Injection Moulding': 'Шприцоване', 'Logistics': 'Спедиция',
    'Machine Maintenance': 'Поддръжка машини', 'Machining': 'Механична обработка',
    'Press Shop': 'Преси', 'Production': 'Производство',
    'Production Blanks': 'Заготовки за производство', 'Project Management': 'Управление на проекти',
    'Purchasing': 'Покупки', 'Quality Assurance': 'Осигуряване на качеството',
    'Quality Management': 'Управление на качеството', 'R&D': 'Развойна дейност',
    'Sales': 'Продажби', 'Tool Design': 'Дизайн на инструменти',
    'Tooling Dept.': 'Инструментален отдел', 'Transport': 'ТРАНСПОРТ',
    'Warehouse': 'Склад', 'Welding': 'Заваръчно',
}

def validate_gender(val):
    if not val: return None
    v = str(val).strip().upper()
    if v not in ("M","F","D"):
        raise ValueError(f"Ungültiger Gender-Wert '{val}' — erwartet M, F oder D")
    return v

def validate_date(val):
    if val is None or str(val).strip() == "": return None
    if isinstance(val, (datetime, date)):
        return (val.date() if isinstance(val, datetime) else val).isoformat()
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try: return datetime.strptime(val, fmt).date().isoformat()
        except ValueError: pass
    raise ValueError(f"Ungültiges Datum '{val}' — erwartet DD.MM.YYYY")

def validate_fte(val):
    if val is None or str(val).strip() == "": return None
    try: return float(str(val).replace(",","."))
    except: raise ValueError(f"Ungültiger FTE-Wert: '{val}'")

def validate_net(val):
    if val is None or str(val).strip() == "": return None
    try: return float(str(val).replace(",",".").replace("'",""))
    except: raise ValueError(f"Ungültiger Gehaltswert: '{val}'")

def clean_erp(val):
    if val is None: return ""
    s = str(val).strip()
    if s.endswith(".0"): s = s[:-2]
    return s

def read_sheet(wb, name_candidates):
    for name in name_candidates:
        for sn in wb.sheetnames:
            if name.lower() in sn.lower():
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: return [], []
                headers = [str(h).strip().lower().replace(" ","_") if h else "" for h in rows[0]]
                data = []
                for row in rows[1:]:
                    if all(v is None or str(v).strip() == "" for v in row): continue
                    data.append(dict(zip(headers, row)))
                return headers, data
    return [], []

def parse_ma_rows(raw_rows):
    rows, errors = [], []
    for i, row in enumerate(raw_rows, 2):
        erp_id = clean_erp(row.get("erp_id"))
        if not erp_id: continue
        try:
            dept_en = str(row.get("department_en") or "").strip() or None
            dept_bg = DEPT_BG_BY_EN.get(dept_en) if dept_en else None
            status = str(row.get("status") or "").strip() or None
            if status and status not in ("Active","Inactive"):
                raise ValueError(f"Status muss 'Active' oder 'Inactive' sein, nicht '{status}'")
            emp_type = str(row.get("employee_type") or "").strip() or None
            if emp_type and emp_type not in ("Regular", "Trainee", "Apprentice"):
                raise ValueError(f"employee_type muss Regular, Trainee oder Apprentice sein, nicht '{emp_type}'")
            rows.append({
                "erp_id":        erp_id,
                "first_name":    str(row.get("first_name") or "").strip() or None,
                "last_name":     str(row.get("last_name") or "").strip() or None,
                "gender":        validate_gender(row.get("gender")),
                "birth_date":    validate_date(row.get("birth_date")),
                "department_en": dept_en,
                "department":    dept_bg,
                "cost_center":   str(row.get("cost_center") or "").strip() or None,
                "position_en":   str(row.get("position_en") or "").strip() or None,
                "entry_date":    validate_date(row.get("entry_date")),
                "fte":           validate_fte(row.get("fte")),
                "status":        status,
                "employee_type": emp_type,
            })
        except ValueError as e:
            errors.append(f"Mitarbeiter Zeile {i} (ERP {erp_id}): {e}")
    return rows, errors

def parse_lohn_rows(raw_rows):
    rows, errors = [], []
    for i, row in enumerate(raw_rows, 2):
        erp_id = clean_erp(row.get("erp_id"))
        if not erp_id: continue
        try:
            rows.append({
                "erp_id":    erp_id,
                "band_code": str(row.get("band_code") or "").strip() or None,
                "surcharge": str(row.get("surcharge") or "").strip().upper() or None,
                "actual_net":validate_net(row.get("actual_net")),
            })
        except ValueError as e:
            errors.append(f"Lohn Zeile {i} (ERP {erp_id}): {e}")
    return rows, errors

def upsert_employee(cur, r):
    cur.execute("SELECT id FROM employees WHERE erp_id=?", (r["erp_id"],))
    existing = cur.fetchone()
    fields = {k:v for k,v in r.items() if k != "erp_id" and v is not None}
    if existing:
        if not fields: return "skip"
        set_clause = ", ".join(f"{k}=?" for k in fields)
        cur.execute(f"UPDATE employees SET {set_clause} WHERE erp_id=?",
                    list(fields.values()) + [r["erp_id"]])
        return "updated"
    else:
        # New employee — department is required
        dept = r.get("department") or r.get("department_en") or "Unknown"
        all_fields = {"erp_id": r["erp_id"], "department": dept, **fields}
        cols = ", ".join(all_fields.keys())
        placeholders = ", ".join("?" for _ in all_fields)
        cur.execute(f"INSERT INTO employees ({cols}) VALUES ({placeholders})",
                    list(all_fields.values()))
        return "inserted"

def upsert_band(cur, r):
    cur.execute("SELECT id FROM employee_band_positions WHERE erp_id=?", (r["erp_id"],))
    existing = cur.fetchone()
    fields = {k:v for k,v in r.items() if k != "erp_id" and v is not None}
    if existing:
        if not fields: return "skip"
        set_clause = ", ".join(f"{k}=?" for k in fields)
        cur.execute(f"UPDATE employee_band_positions SET {set_clause} WHERE erp_id=?",
                    list(fields.values()) + [r["erp_id"]])
        return "updated"
    else:
        all_fields = {"erp_id": r["erp_id"], **fields}
        if not fields: return "skip"
        cols = ", ".join(all_fields.keys())
        placeholders = ", ".join("?" for _ in all_fields)
        cur.execute(f"INSERT INTO employee_band_positions ({cols}) VALUES ({placeholders})",
                    list(all_fields.values()))
        return "inserted"

def main():
    dry_run = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("Verwendung: python import_names.py import_template.xlsx [--dry-run]")
        sys.exit(1)

    path = args[0]
    if not os.path.exists(path):
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"Datenbank nicht gefunden: {DB_PATH}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Fehlt: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)

    _, ma_raw  = read_sheet(wb, ["1_mitarbeiter","mitarbeiter","employees"])
    _, lohn_raw = read_sheet(wb, ["2_lohndaten","lohndaten","salary","lohn"])

    ma_rows,   ma_errors   = parse_ma_rows(ma_raw)
    lohn_rows, lohn_errors = parse_lohn_rows(lohn_raw)

    all_errors = ma_errors + lohn_errors
    if all_errors:
        print("Validierungsfehler:")
        for e in all_errors: print(f"  {e}")
        sys.exit(1)

    print(f"Mitarbeiter-Zeilen: {len(ma_rows)}")
    print(f"Lohn-Zeilen:        {len(lohn_rows)}")

    if dry_run:
        print("\n── DRY RUN ────────────────────────────────────")
        for r in ma_rows[:5]:
            print(f"  MA  ERP {r['erp_id']:>6}  {r.get('first_name') or '–':14} {r.get('last_name') or '–':18}  {r.get('gender') or '–'}  {r.get('status') or '–'}")
        for r in lohn_rows[:5]:
            print(f"  LOHN ERP {r['erp_id']:>6}  Band: {r.get('band_code') or '–':10}  Sur: {r.get('surcharge') or '–'}  Net: {r.get('actual_net') or '–'}")
        print("────────────────────────────────────────────────")
        return

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    stats = {"ma_updated":0,"ma_inserted":0,"lohn_updated":0,"lohn_inserted":0}

    for r in ma_rows:
        res = upsert_employee(cur, r)
        if res == "updated":  stats["ma_updated"]  += 1
        if res == "inserted": stats["ma_inserted"] += 1

    for r in lohn_rows:
        res = upsert_band(cur, r)
        if res == "updated":  stats["lohn_updated"]  += 1
        if res == "inserted": stats["lohn_inserted"] += 1

    conn.commit()
    conn.close()

    print(f"\nMitarbeiter  — aktualisiert: {stats['ma_updated']:3}  neu: {stats['ma_inserted']:3}")
    print(f"Lohndaten    — aktualisiert: {stats['lohn_updated']:3}  neu: {stats['lohn_inserted']:3}")
    print("Fertig. Backend neu starten damit Änderungen sichtbar werden.")

if __name__ == "__main__":
    main()
