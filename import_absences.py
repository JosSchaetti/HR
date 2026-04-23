"""
import_absences.py — Absenzen in einen monatlichen Review-Eintrag importieren
==============================================================================
Erwartet eine .xlsx-Datei mit einem Sheet "Absenzen" (oder "absences"):

    erp_id* | absence_days | absence_type
    --------|--------------|-------------------------------------
    176     | 2.5          | sick
    3       | 5            | vacation
    88      | 0.5          | unpaid

absence_type: sick | vacation | unpaid | other  (leer = other)

Verwendung:
    python import_absences.py absenzen_2026_04.xlsx --period 2026-04
    python import_absences.py absenzen_2026_04.xlsx --period 2026-04 --dry-run

Der Review fuer die Periode wird automatisch angelegt falls er noch nicht existiert.
"""

import sys, os, sqlite3
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "glatec_hr.db")
VALID_TYPES = {"sick", "vacation", "unpaid", "other"}


def clean_erp(val):
    if val is None: return ""
    s = str(val).strip()
    if s.endswith(".0"): s = s[:-2]
    return s

def validate_days(val):
    if val is None or str(val).strip() == "": return 0.0
    try: return float(str(val).replace(",", "."))
    except: raise ValueError(f"Ungueltige Tageszahl: \'{val}\'")

def validate_type(val):
    if not val or str(val).strip() == "": return "other"
    v = str(val).strip().lower()
    if v not in VALID_TYPES:
        raise ValueError(f"absence_type muss sick/vacation/unpaid/other sein, nicht \'{val}\'")
    return v

def read_sheet(wb):
    for name in ["absenzen", "absences", "absence"]:
        for sn in wb.sheetnames:
            if name in sn.lower():
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: return [], []
                headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
                data = []
                for row in rows[1:]:
                    if all(v is None or str(v).strip() == "" for v in row): continue
                    data.append(dict(zip(headers, row)))
                return headers, data
    return [], []


def main():
    dry_run = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    period = None
    for i, a in enumerate(sys.argv[1:]):
        if a == "--period" and i+2 <= len(sys.argv)-1:
            period = sys.argv[i+2]

    if not args:
        print("Verwendung: python import_absences.py datei.xlsx --period YYYY-MM [--dry-run]")
        sys.exit(1)

    if not period:
        print("Fehler: --period YYYY-MM fehlt")
        sys.exit(1)

    path = args[0]
    if not os.path.exists(path):
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"Datenbank nicht gefunden: {DB_PATH}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Fehlt: pip install openpyxl"); sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    _, raw = read_sheet(wb)
    if not raw:
        print("Kein Sheet 'Absenzen' gefunden. Erwartet: Sheet-Name enthaelt 'absenzen' oder 'absences'.")
        sys.exit(1)

    rows, errors = [], []
    for i, row in enumerate(raw, 2):
        erp_id = clean_erp(row.get("erp_id"))
        if not erp_id: continue
        try:
            rows.append({
                "erp_id": erp_id,
                "absence_days": validate_days(row.get("absence_days")),
                "absence_type": validate_type(row.get("absence_type")),
            })
        except ValueError as e:
            errors.append(f"Zeile {i}: {e}")

    if errors:
        print("Validierungsfehler:")
        for e in errors: print(f"  {e}")
        sys.exit(1)

    print(f"Periode: {period}")
    print(f"Absenz-Zeilen: {len(rows)}")

    if dry_run:
        print("\n── DRY RUN ─────────────────────────────────────────────────")
        for r in rows[:5]:
            print(f"  ERP {r['erp_id']:>6}  {r['absence_days']:5.1f} Tage  [{r['absence_type']}]")
        print("────────────────────────────────────────────────────────────")
        return

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Ensure review exists
    cur.execute("SELECT id, status FROM monthly_reviews WHERE period=?", (period,))
    rev = cur.fetchone()
    if rev:
        review_id, status = rev
        if status == "approved":
            print(f"Fehler: Review {period} ist bereits genehmigt. Zuerst zuruecksetzen.")
            conn.close(); sys.exit(1)
        print(f"Review {period} gefunden (id={review_id}, status={status})")
    else:
        cur.execute(
            "INSERT INTO monthly_reviews (period, status, created_by) VALUES (?, \'draft\', \'import_script\')",
            (period,)
        )
        review_id = cur.lastrowid
        print(f"Review {period} angelegt (id={review_id})")

    stats = {"updated": 0, "inserted": 0}
    for r in rows:
        cur.execute(
            "SELECT id FROM monthly_review_entries WHERE review_id=? AND erp_id=?",
            (review_id, r["erp_id"])
        )
        existing = cur.fetchone()
        now = datetime.utcnow().isoformat()
        if existing:
            cur.execute(
                """UPDATE monthly_review_entries
                   SET absence_days=?, absence_type=?, updated_by='import_script', updated_at=?
                   WHERE review_id=? AND erp_id=?""",
                (r["absence_days"], r["absence_type"], now, review_id, r["erp_id"])
            )
            stats["updated"] += 1
        else:
            cur.execute(
                """INSERT INTO monthly_review_entries
                   (review_id, erp_id, absence_days, absence_type, bonus_amount, updated_by, updated_at)
                   VALUES (?, ?, ?, ?, 0, 'import_script', ?)""",
                (review_id, r["erp_id"], r["absence_days"], r["absence_type"], now)
            )
            stats["inserted"] += 1

    conn.commit()
    conn.close()
    print(f"\nAbsenzen  — aktualisiert: {stats['updated']:3}  neu: {stats['inserted']:3}")
    print(f"Review ID {review_id} — Backend neu starten damit Aenderungen sichtbar werden.")


if __name__ == "__main__":
    main()
