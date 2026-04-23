"""
import_qualifications.py — Qualifikationsmatrix aus Excel importieren
=====================================================================
Erwartet Qualification_Matrix_v02.xlsx (oder kompatibles Format)

Sheet "Qualification Matrix":
  Row 3: Abteilungsgruppen (Injection molding, Press, Assembly, Welding, Mechanical)
  Row 4: WC-Codes (WC 1822.10, ...)
  Row 5: Maschinenbezeichnungen
  Row 6+: Mitarbeiter (Spalte D = Employee ID / ERP ID, Spalten H-BB = Qualifikation 1-4)

Ausführen:
    python import_qualifications.py Qualification_Matrix_v02.xlsx
    python import_qualifications.py Qualification_Matrix_v02.xlsx --dry-run
"""

import sys, os, sqlite3

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "glatec_hr.db")

WC_START_COL = 7   # 0-indexed, column H
WC_END_COL   = 53  # 0-indexed, column BB

def main():
    dry_run = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("Verwendung: python import_qualifications.py Qualification_Matrix.xlsx [--dry-run]")
        sys.exit(1)

    path = args[0]
    if not os.path.exists(path):
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Fehlt: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Qualification Matrix" not in wb.sheetnames:
        print("Sheet 'Qualification Matrix' nicht gefunden")
        sys.exit(1)

    ws = wb["Qualification Matrix"]
    rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
    row3, row4, row5 = rows[2], rows[3], rows[4]

    # Build WC list
    dept_group = None
    wc_list = []
    for i in range(WC_START_COL, WC_END_COL + 1):
        if row3[i]: dept_group = str(row3[i]).strip()
        wc_code = row4[i]
        machine = row5[i]
        if wc_code:
            wc_list.append({
                "col": i,
                "code": str(wc_code).strip(),
                "machine": str(machine).strip() if machine else None,
                "dept_group": dept_group,
            })

    print(f"Work Centers gefunden: {len(wc_list)}")

    # Build employee rows
    data_rows = []
    for r in rows[5:]:
        raw_id = r[3]
        if raw_id is None: continue
        erp_id = str(int(float(raw_id))) if isinstance(raw_id, (int, float)) else str(raw_id).strip()
        if not erp_id or erp_id.lower() == "none": continue
        scores = {}
        for wc in wc_list:
            val = r[wc["col"]]
            if val is not None:
                try:
                    level = int(float(val))
                    if 1 <= level <= 4:
                        scores[wc["code"]] = level
                except (ValueError, TypeError):
                    pass
        data_rows.append({"erp_id": erp_id, "scores": scores})

    print(f"Mitarbeiter-Zeilen: {len(data_rows)}")
    total_scores = sum(len(r["scores"]) for r in data_rows)
    print(f"Qualifikationseinträge (1-4): {total_scores}")

    if dry_run:
        print("\n── DRY RUN ──────────────────────────────────────")
        for r in data_rows[:5]:
            print(f"  ERP {r['erp_id']:>6}  Scores: {dict(list(r['scores'].items())[:4])}")
        print("─────────────────────────────────────────────────")
        return

    if not os.path.exists(DB_PATH):
        print(f"Datenbank nicht gefunden: {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    # Upsert work centers
    for wc in wc_list:
        cur.execute("""
            INSERT INTO work_centers (code, machine, dept_group, col_index)
            VALUES (?,?,?,?)
            ON CONFLICT(code) DO UPDATE SET machine=excluded.machine, dept_group=excluded.dept_group
        """, (wc["code"], wc["machine"], wc["dept_group"], wc["col"]))

    # Upsert qualifications
    inserted = updated = 0
    for row in data_rows:
        for wc_code, level in row["scores"].items():
            cur.execute("SELECT id FROM qualifications WHERE erp_id=? AND wc_code=?",
                        (row["erp_id"], wc_code))
            if cur.fetchone():
                cur.execute("UPDATE qualifications SET level=? WHERE erp_id=? AND wc_code=?",
                            (level, row["erp_id"], wc_code))
                updated += 1
            else:
                cur.execute("INSERT INTO qualifications (erp_id, wc_code, level) VALUES (?,?,?)",
                            (row["erp_id"], wc_code, level))
                inserted += 1

    conn.commit()
    conn.close()
    print(f"\nWork Centers upserted: {len(wc_list)}")
    print(f"Qualifikationen — neu: {inserted}  aktualisiert: {updated}")
    print("Fertig. Backend neu starten.")

if __name__ == "__main__":
    main()
