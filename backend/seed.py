"""
Glatec HR Platform – Database Seed Script
Lädt alle Daten aus den Excel-Files und erstellt Initial-User.
Pfad zu den Excel-Files anpassen falls nötig.
"""
import sys
import os
import openpyxl
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import engine, SessionLocal
import models
from auth import hash_password

# ── Pfad zu den Excel-Quelldateien ──────────────────────────────────────────
# Standardmässig wird erwartet dass die xlsx-Files im selben Ordner liegen.
# Passe die Pfade hier an falls die Files woanders sind.
EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")

WAGBAND_FILE = os.path.join(EXCEL_DIR, "Glatec_WageBand_C_260310.xlsx")
PAYROLL_FILE = os.path.join(EXCEL_DIR, "Payroll_Template.xlsx")

# Fallback: direkt im Backend-Ordner suchen
if not os.path.exists(WAGBAND_FILE):
    WAGBAND_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Glatec_WageBand_C_260310.xlsx")
if not os.path.exists(PAYROLL_FILE):
    PAYROLL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Payroll_Template.xlsx")


def create_tables():
    models.Base.metadata.create_all(bind=engine)
    print("✓ Tabellen erstellt")


def seed_users(db):
    if db.query(models.User).count() > 0:
        print("  Users bereits vorhanden – übersprungen")
        return
    users = [
        models.User(
            username="jos",
            full_name="Jos Schaetti",
            role="admin",
            hashed_password=hash_password("Admin2026"),
            is_active=True
        ),
        models.User(
            username="maria",
            full_name="Maria",
            role="manager",
            hashed_password=hash_password("Manager2026"),
            is_active=True
        ),
        models.User(
            username="hr",
            full_name="HR Team",
            role="hr",
            hashed_password=hash_password("HR2026"),
            is_active=True
        ),
        models.User(
            username="lohn",
            full_name="Lohnbuchhaltung",
            role="payroll",
            hashed_password=hash_password("Payroll2026"),
            is_active=True
        ),
    ]
    db.add_all(users)
    db.commit()
    print("✓ Users erstellt")
    print("  jos        / Admin2026     (Admin)")
    print("  maria      / Manager2026   (Manager)")
    print("  hr         / HR2026        (HR)")
    print("  lohn       / Payroll2026   (Lohnbuchhaltung)")
    print("  WICHTIG: Passwörter nach erstem Login in der Benutzerverwaltung ändern!")


def safe_float(val):
    try:
        if val is None or val == '–' or val == '' or val == '#N/A':
            return None
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' and s != '–' else None


def seed_wage_bands(db):
    if not os.path.exists(WAGBAND_FILE):
        print(f"  ⚠ WageBand-File nicht gefunden: {WAGBAND_FILE}")
        print("    Bitte Excel-Files in den 'data' Ordner kopieren")
        seed_wage_bands_static(db)
        return

    wb = openpyxl.load_workbook(WAGBAND_FILE, data_only=True)
    ws = wb["HR Lookup"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        dept, code, en_title, surcharge, min_net, target_net, max_net, notes = row[:8]
        if not code or not en_title:
            continue
        existing = db.query(models.WageBand).filter(models.WageBand.code == str(code)).first()
        if existing:
            continue
        band = models.WageBand(
            code=str(code),
            department=str(dept),
            en_title=str(en_title),
            surcharge=str(surcharge) if surcharge else "C",
            min_net=safe_float(min_net) or 0,
            target_net=safe_float(target_net) or 0,
            max_net=safe_float(max_net) or 0,
            notes=safe_str(notes)
        )
        db.add(band)
        count += 1
    db.commit()
    print(f"✓ Wage Bands geladen: {count} Einträge")

    # Clear existing band positions to prevent duplicates on re-run
    existing_pos = db.query(models.EmployeeBandPosition).count()
    if existing_pos > 0:
        db.query(models.EmployeeBandPosition).delete()
        db.commit()
        print(f"  (employee_band_positions: {existing_pos} alte Einträge gelöscht)")

    # Lade Band-Positionen aus den Abteilungs-Sheets
    sections = {
        "Production": ["Welding + Injection", "Machining", "Toolroom", "Assembly", "Pressing"],
        "Engineering": ["Engineering / Automation"],
        "Quality": ["Quality"],
        "Admin": ["Warehouse", "Shipping", "Transport", "Administration"]
    }
    sheet_names = ["Production", "Engineering", "Quality", "Admin"]
    pos_count = 0
    current_section = None

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws_dept = wb[sheet_name]
        current_section = None
        for row in ws_dept.iter_rows(min_row=4, values_only=True):
            if not any(v for v in row):
                continue
            # Detect section header rows (e.g. "  ▶  Welding + Injection")
            first_cell = str(row[0]) if row[0] else ""
            if "▶" in first_cell or (row[0] and not isinstance(row[0], (int, float))):
                potential_section = first_cell.strip().replace("▶", "").strip()
                if potential_section and len(potential_section) > 1:
                    current_section = potential_section
                    continue

            num, erp_id, bg_name, en_name, pos_bg, band_code, en_title, surcharge, min_n, target_n, max_n, actual_n, band_pos, bar, cc = (list(row) + [None]*15)[:15]

            if not isinstance(num, (int, float)):
                continue
            if not erp_id:
                continue

            pos = models.EmployeeBandPosition(
                erp_id=safe_str(erp_id),
                section=sheet_name,
                department_group=current_section,
                position_bg=safe_str(pos_bg),
                band_code=safe_str(band_code),
                en_title=safe_str(en_title),
                surcharge=safe_str(surcharge),
                min_net=safe_float(min_n),
                target_net=safe_float(target_n),
                max_net=safe_float(max_n),
                actual_net=safe_float(actual_n),
                band_position=safe_str(band_pos),
                cost_center=safe_str(cc)
            )
            db.add(pos)
            pos_count += 1

    db.commit()
    print(f"✓ Band-Positionen geladen: {pos_count} Mitarbeiter")


def seed_wage_bands_static(db):
    """Fallback: statische Wage Band Daten wenn kein Excel vorhanden"""
    bands = [
        ("H1_AM", "Assembly", "Assembly Helper Gr.1", "C", 496, 496, 560),
        ("H2_AM", "Assembly", "Assembly Helper Gr.2", "C", 496, 600, 700),
        ("O1_AM", "Assembly", "Assembly Operator Gr.1", "C", 536, 650, 750),
        ("O2_AM", "Assembly", "Assembly Operator Gr.2", "C", 600, 716, 870),
        ("E_AM", "Assembly", "Setter – Assembly", "C", 797, 1028, 1350),
        ("G_AM", "Assembly", "Section Leader – Assembly", "C", 736, 948, 1500),
        ("O1_WE", "Welding", "Welding Operator Gr.1", "A", 632, 825, 957),
        ("O2_WE", "Welding", "Welding Operator Gr.2", "A", 546, 923, 1078),
        ("G_WE", "Welding", "Section Leader – Welding", "A", 610, 788, 957),
        ("O1_MC", "Machining", "Machining Operator Gr.1", "B", 604, 735, 861),
        ("E1_MC", "Machining", "Setter 1 – Machining (CNC)", "B", 804, 1036, 1208),
        ("G_MC", "Machining", "Section Leader – Machining", "B", 1125, 1449, 1680),
        ("QA1", "Quality", "Quality Inspector", "C", 615, 793, 1023),
        ("QA4", "Quality", "Quality Engineer", "C", 1474, 1900, 2200),
        ("QA5", "Quality", "Quality Manager", "C", 1832, 2326, 2700),
        ("PM2", "Engineering", "Head of Production Technology", "C", 1432, 1846, 2200),
        ("AE8", "Engineering", "Automation Engineer", "C", 1327, 1710, 2000),
        ("MG2", "Admin", "Production Manager", "C", 1627, 2096, 2500),
        ("MG3", "Admin", "Production Director", "C", 1740, 2224, 2800),
        ("HR2", "Admin", "HR Officer", "C", 913, 1176, 1400),
        ("FI2", "Admin", "Senior Accountant", "C", 885, 1140, 1400),
    ]
    for code, dept, title, sur, mn, tg, mx in bands:
        if not db.query(models.WageBand).filter(models.WageBand.code == code).first():
            db.add(models.WageBand(code=code, department=dept, en_title=title,
                                   surcharge=sur, min_net=mn, target_net=tg, max_net=mx))
    db.commit()
    print("✓ Wage Bands (statisch) geladen")


def seed_employees(db):
    if not os.path.exists(PAYROLL_FILE):
        print(f"  ⚠ Payroll-File nicht gefunden: {PAYROLL_FILE}")
        return

    wb = openpyxl.load_workbook(PAYROLL_FILE, data_only=True)
    ws = wb["Master Data"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_id, name, dept, pos, entry_date, fte, cc, status = (list(row) + [None]*8)[:8]
        if not emp_id or not isinstance(emp_id, (int, float)):
            continue
        existing = db.query(models.Employee).filter(models.Employee.erp_id == str(int(emp_id))).first()
        if existing:
            continue
        entry = None
        if isinstance(entry_date, datetime):
            entry = entry_date.date()
        elif isinstance(entry_date, date):
            entry = entry_date

        emp = models.Employee(
            erp_id=str(int(emp_id)),
            department=safe_str(dept) or "Unknown",
            position_bg=safe_str(pos),
            entry_date=entry,
            fte=safe_float(fte) or 1.0,
            cost_center=str(int(cc)) if cc and isinstance(cc, (int, float)) else safe_str(cc),
            status=safe_str(status) or "Active"
        )
        db.add(emp)
        count += 1
    db.commit()
    print(f"✓ Employees geladen: {count} Mitarbeiter")


def seed_salary(db):
    if not os.path.exists(PAYROLL_FILE):
        return

    # Clear existing data to prevent duplicates on re-run
    existing = db.query(models.SalaryRecord).count()
    if existing > 0:
        db.query(models.SalaryRecord).delete()
        db.commit()
        print(f"  (salary_records: {existing} alte Einträge gelöscht)")

    wb = openpyxl.load_workbook(PAYROLL_FILE, data_only=True)
    ws = wb["Salary History"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_id, eff_date, name, cc, dept, base_g, prof_g, bonus, food, total_g, netto, total_mn, annual_n, month_ec, annual_ec = (list(row) + [None]*15)[:15]
        if not emp_id or not isinstance(emp_id, (int, float)):
            continue

        eff = None
        if isinstance(eff_date, datetime):
            eff = eff_date.date()

        rec = models.SalaryRecord(
            employee_id=int(emp_id),
            erp_id=str(int(emp_id)),
            effective_date=eff,
            department=safe_str(dept),
            cost_center=str(int(cc)) if cc and isinstance(cc, (int, float)) and str(cc) != '#N/A' else None,
            base_salary_gross=safe_float(base_g),
            prof_exp_gross=safe_float(prof_g),
            fixed_bonus=safe_float(bonus),
            food_voucher=safe_float(food),
            total_gross=safe_float(total_g),
            salary_netto=safe_float(netto),
            total_monthly_netto=safe_float(total_mn),
            annual_netto=safe_float(annual_n),
            month_employer_cost=safe_float(month_ec),
            annual_employer_cost=safe_float(annual_ec)
        )
        db.add(rec)
        count += 1
    db.commit()
    print(f"✓ Salary Records geladen: {count} Einträge")


def seed_budget(db):
    if not os.path.exists(PAYROLL_FILE):
        return

    # Clear existing data to prevent duplicates on re-run
    existing = db.query(models.BudgetControl).count()
    if existing > 0:
        db.query(models.BudgetControl).delete()
        db.commit()
        print(f"  (budget_control: {existing} alte Einträge gelöscht)")

    wb = openpyxl.load_workbook(PAYROLL_FILE, data_only=True)
    ws = wb["Monthly Budget Control"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        month, dept, cc, ann_budget, mon_budget, actual, variance, var_pct = (list(row) + [None]*8)[:8]
        if not month or not isinstance(month, (int, float)):
            continue

        bc = models.BudgetControl(
            month=int(month),
            department=safe_str(dept) or "Unknown",
            cost_center=int(cc) if cc and isinstance(cc, (int, float)) else None,
            approved_annual_budget=safe_float(ann_budget),
            approved_monthly_budget=safe_float(mon_budget),
            actual_payroll=safe_float(actual),
            variance=safe_float(variance),
            variance_pct=safe_float(var_pct)
        )
        db.add(bc)
        count += 1
    db.commit()
    print(f"✓ Budget Control geladen: {count} Abteilungen")


if __name__ == "__main__":
    print("\n=== Glatec HR Platform – Datenbank initialisieren ===\n")

    # Excel-Files in data-Ordner kopieren falls nötig
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
    os.makedirs(data_dir, exist_ok=True)

    # Suche Excel-Files auch im HR-Quellordner
    possible_paths = [
        os.path.expanduser("~/Downloads"),
        os.path.dirname(os.path.abspath(__file__)),
        data_dir
    ]

    create_tables()

    db = SessionLocal()
    try:
        seed_users(db)
        seed_wage_bands(db)
        seed_employees(db)
        seed_salary(db)
        seed_budget(db)
        print('\n=== Initialisierung abgeschlossen ===')
    finally:
        db.close()
